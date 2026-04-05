"""
Tests for the Excel parser service.
"""

import pytest
from unittest.mock import patch

from app.services.excel_parser import parse_excel, _normalize_cell, _format_cell

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def simple_xlsx(tmp_path) -> str:
    """Create a minimal XLSX file for testing."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Auxiliar IVA"
    ws.append(["Fecha", "Cuenta", "Detalle", "Débito", "Crédito", "Saldo"])
    ws.append(["2026-01-15", "240802", "IVA Descontable", 285000, 0, 285000])
    ws.append(["2026-01-20", "240808", "IVA Generado", 0, 190000, 95000])
    ws.append([None, None, None, None, None, None])  # empty row
    ws.append(["2026-01-31", "240802", "Ajuste IVA", 50000, 0, 145000])

    # Add second sheet
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Concepto", "Valor"])
    ws2.append(["Total IVA Descontable", 335000])
    ws2.append(["Total IVA Generado", 190000])

    path = tmp_path / "test_auxiliar.xlsx"
    wb.save(str(path))
    return str(path)


@pytest.fixture
def empty_xlsx(tmp_path) -> str:
    """Create an empty XLSX file."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Empty"
    path = tmp_path / "empty.xlsx"
    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestParseExcel:
    def test_parses_two_sheets(self, simple_xlsx):
        markdown, sheets = parse_excel(simple_xlsx)
        assert len(sheets) == 2
        assert sheets[0]["sheet_name"] == "Auxiliar IVA"
        assert sheets[1]["sheet_name"] == "Resumen"

    def test_headers_extracted(self, simple_xlsx):
        _, sheets = parse_excel(simple_xlsx)
        assert sheets[0]["headers"] == [
            "Fecha",
            "Cuenta",
            "Detalle",
            "Débito",
            "Crédito",
            "Saldo",
        ]

    def test_rows_extracted_skipping_empty(self, simple_xlsx):
        _, sheets = parse_excel(simple_xlsx)
        # 3 data rows (empty row skipped)
        assert len(sheets[0]["rows"]) == 3

    def test_row_values_correct(self, simple_xlsx):
        _, sheets = parse_excel(simple_xlsx)
        first_row = sheets[0]["rows"][0]
        assert first_row["Fecha"] == "2026-01-15"
        assert str(first_row["Cuenta"]) == "240802"
        assert first_row["Débito"] == 285000

    def test_markdown_contains_table(self, simple_xlsx):
        markdown, _ = parse_excel(simple_xlsx)
        assert "## Auxiliar IVA" in markdown
        assert "| Fecha |" in markdown
        assert "240802" in markdown

    def test_markdown_contains_both_sheets(self, simple_xlsx):
        markdown, _ = parse_excel(simple_xlsx)
        assert "## Auxiliar IVA" in markdown
        assert "## Resumen" in markdown

    def test_second_sheet_data(self, simple_xlsx):
        _, sheets = parse_excel(simple_xlsx)
        resumen = sheets[1]
        assert len(resumen["rows"]) == 2
        assert resumen["rows"][0]["Concepto"] == "Total IVA Descontable"
        assert resumen["rows"][0]["Valor"] == 335000

    def test_empty_xlsx_returns_empty(self, empty_xlsx):
        markdown, sheets = parse_excel(empty_xlsx)
        assert sheets == []
        assert markdown == ""

    def test_file_not_found_raises(self):
        with pytest.raises(FileNotFoundError):
            parse_excel("/nonexistent/file.xlsx")

    def test_openpyxl_not_installed(self):
        with patch("app.services.excel_parser.load_workbook", None):
            with pytest.raises(RuntimeError, match="openpyxl"):
                parse_excel("/any/file.xlsx")


class TestNormalizeCell:
    def test_none_returns_none(self):
        assert _normalize_cell(None) is None

    def test_integer_float_becomes_int(self):
        assert _normalize_cell(100.0) == 100

    def test_fractional_float_stays(self):
        assert _normalize_cell(99.5) == 99.5

    def test_whitespace_string_becomes_none(self):
        assert _normalize_cell("   ") is None

    def test_string_stripped(self):
        assert _normalize_cell("  hello  ") == "hello"


class TestFormatCell:
    def test_none_returns_empty(self):
        assert _format_cell(None) == ""

    def test_integer_formatted_with_commas(self):
        assert _format_cell(1500000) == "1,500,000"

    def test_float_formatted_with_decimals(self):
        assert _format_cell(1234.56) == "1,234.56"

    def test_pipe_escaped(self):
        assert _format_cell("a|b") == "a\\|b"
