"""
Excel parser service for XLSX document ingestion.

Converts multi-sheet XLSX files into:
  1. Markdown text representation (for LLM interpretation, similar to LlamaParse output)
  2. Structured tabular data (list of sheet dicts with headers and rows)
"""

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    load_workbook = None  # type: ignore[assignment,misc]
    get_column_letter = None  # type: ignore[assignment,misc]


def parse_excel(file_path: str) -> tuple[str, list[dict[str, Any]]]:
    """
    Parse an XLSX file into markdown text and structured tabular data.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        (markdown_text, sheets_data) where:
          - markdown_text: all sheets as markdown tables (for LLM)
          - sheets_data: [{sheet_name, headers, rows: [dict]}]

    Raises:
        RuntimeError: If openpyxl is not installed.
        FileNotFoundError: If the file doesn't exist.
    """
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is required for Excel parsing. "
            "Install it: pip install openpyxl>=3.1.0"
        )

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheets_data: list[dict[str, Any]] = []
    markdown_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_raw: list[list[Any]] = []

        for row in ws.iter_rows(values_only=True):
            rows_raw.append([_normalize_cell(cell) for cell in row])

        if not rows_raw:
            continue

        # Find headers: first non-empty row
        header_idx = 0
        for i, row in enumerate(rows_raw):
            if any(cell for cell in row):
                header_idx = i
                break

        headers = [str(cell) if cell else f"col_{j}" for j, cell in enumerate(rows_raw[header_idx])]
        data_rows: list[dict[str, Any]] = []

        for row in rows_raw[header_idx + 1:]:
            if not any(cell for cell in row):
                continue  # skip empty rows
            row_dict = {}
            for j, cell in enumerate(row):
                key = headers[j] if j < len(headers) else f"col_{j}"
                row_dict[key] = cell
            data_rows.append(row_dict)

        sheets_data.append({
            "sheet_name": sheet_name,
            "headers": headers,
            "rows": data_rows,
        })

        # Build markdown table for this sheet
        md = _build_markdown_table(sheet_name, headers, data_rows)
        markdown_parts.append(md)

    wb.close()

    markdown_text = "\n\n".join(markdown_parts)
    logger.info(
        "excel_parser: parsed %d sheets, %d total rows from %s",
        len(sheets_data),
        sum(len(s["rows"]) for s in sheets_data),
        path.name,
    )
    return markdown_text, sheets_data


def _normalize_cell(value: Any) -> Any:
    """Normalize a cell value for consistent output."""
    if value is None:
        return None
    if isinstance(value, float):
        # Remove trailing .0 for integer-like floats
        if value == int(value):
            return int(value)
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else None
    return value


def _build_markdown_table(
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> str:
    """Build a markdown table representation of a sheet."""
    lines: list[str] = []
    lines.append(f"## {sheet_name}")
    lines.append("")

    if not rows:
        lines.append("*(empty sheet)*")
        return "\n".join(lines)

    # Header row
    lines.append("| " + " | ".join(str(h) for h in headers) + " |")
    lines.append("| " + " | ".join("---" for _ in headers) + " |")

    # Data rows
    for row in rows:
        cells = []
        for h in headers:
            val = row.get(h, "")
            cell_str = _format_cell(val)
            cells.append(cell_str)
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def _format_cell(value: Any) -> str:
    """Format a cell value for markdown display."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # Colombian number formatting: use comma for thousands
        if isinstance(value, float) and value != int(value):
            return f"{value:,.2f}"
        return f"{int(value):,}"
    return str(value).replace("|", "\\|")  # escape pipe in markdown tables
