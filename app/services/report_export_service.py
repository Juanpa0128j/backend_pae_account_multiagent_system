"""
Export financial reports to PDF and Excel formats.

Provides presentation templates for:
- Balance Sheet (Balance General)
- Profit & Loss (Estado de Resultados)
- Cash Flow (Flujo de Caja)
"""

import io
from html import escape
from datetime import datetime
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _format_currency(value: float) -> str:
    """Format a number as Colombian currency."""
    if value is None:
        return "$ 0"
    return f"$ {value:,.0f}"


def _get_cuenta_codigo(cuenta: Dict[str, Any]) -> str:
    """Return account code from normalized or legacy keys."""
    return str(cuenta.get("codigo") or cuenta.get("cuenta") or "N/A")


def _get_cuenta_nombre(cuenta: Dict[str, Any]) -> str:
    """Return account name from normalized or legacy keys."""
    return str(cuenta.get("nombre") or cuenta.get("descripcion") or "N/A")


def _build_table_2cols(data, total_row_color: str):
    table = Table(data, colWidths=[3.5 * inch, 2 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(total_row_color)),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -2),
                    [colors.white, colors.HexColor("#f9f9f9")],
                ),
            ]
        )
    )
    return table


def _build_table_3cols(data, total_row_color: str):
    table = Table(data, colWidths=[1.2 * inch, 2.3 * inch, 2 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(total_row_color)),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -2),
                    [colors.white, colors.HexColor("#f9f9f9")],
                ),
            ]
        )
    )
    return table


def _escape_paragraph_text(value: Any) -> str:
    """Escape dynamic text before interpolation into reportlab Paragraph markup."""
    return escape(str(value))


class BalanceSheetExporter:
    """Export Balance Sheet reports to PDF and Excel."""

    @staticmethod
    def to_pdf(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#1f4788"),
            spaceAfter=6,
            alignment=1,
        )
        subtitle_style = ParagraphStyle(
            "Subtitle",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor("#555555"),
            spaceAfter=12,
            alignment=1,
        )
        section_style = ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontSize=11,
            textColor=colors.HexColor("#1f4788"),
            spaceAfter=6,
            spaceBefore=8,
        )

        story = []
        story.append(Paragraph("BALANCE GENERAL", title_style))
        story.append(Paragraph("Estado de Situacion Financiera", subtitle_style))
        story.append(
            Paragraph(
                f"<b>Empresa:</b> {_escape_paragraph_text(company_name)} | "
                f"<b>Periodo:</b> al {_escape_paragraph_text(report.get('period_end', '--'))} | "
                f"<b>Generado:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.2 * inch))

        activos = float(report.get("activos", 0))
        pasivos = float(report.get("pasivos", 0))
        patrimonio = float(report.get("patrimonio", 0))
        utilidad = float(report.get("utilidad_neta", 0))

        story.append(Paragraph("ACTIVOS (Clase 1)", section_style))
        activos_detalle = report.get("activos_detalle") or []
        if activos_detalle:
            assets_data = [["Cuenta PUC", "Descripcion", "Valor (COP)"]]
            for cuenta in activos_detalle:
                assets_data.append(
                    [
                        _get_cuenta_codigo(cuenta),
                        _get_cuenta_nombre(cuenta),
                        _format_currency(float(cuenta.get("saldo", 0))),
                    ]
                )
            assets_data.append(["", "TOTAL ACTIVOS", _format_currency(activos)])
            story.append(_build_table_3cols(assets_data, "#e8f0f7"))
        else:
            assets_data = [
                ["Concepto", "Valor (COP)"],
                ["Detalle no disponible", ""],
                ["TOTAL ACTIVOS", _format_currency(activos)],
            ]
            story.append(_build_table_2cols(assets_data, "#e8f0f7"))
        story.append(Spacer(1, 0.15 * inch))

        story.append(Paragraph("PASIVOS (Clase 2)", section_style))
        pasivos_detalle = report.get("pasivos_detalle") or []
        if pasivos_detalle:
            liab_data = [["Cuenta PUC", "Descripcion", "Valor (COP)"]]
            for cuenta in pasivos_detalle:
                liab_data.append(
                    [
                        _get_cuenta_codigo(cuenta),
                        _get_cuenta_nombre(cuenta),
                        _format_currency(float(cuenta.get("saldo", 0))),
                    ]
                )
            liab_data.append(["", "TOTAL PASIVOS", _format_currency(pasivos)])
            story.append(_build_table_3cols(liab_data, "#fff4e6"))
        else:
            liab_data = [
                ["Concepto", "Valor (COP)"],
                ["Detalle no disponible", ""],
                ["TOTAL PASIVOS", _format_currency(pasivos)],
            ]
            story.append(_build_table_2cols(liab_data, "#fff4e6"))
        story.append(Spacer(1, 0.15 * inch))

        story.append(Paragraph("PATRIMONIO (Clase 3)", section_style))
        patrimonio_detalle = report.get("patrimonio_detalle") or []
        if patrimonio_detalle:
            equity_data = [["Cuenta PUC", "Descripcion", "Valor (COP)"]]
            for cuenta in patrimonio_detalle:
                equity_data.append(
                    [
                        _get_cuenta_codigo(cuenta),
                        _get_cuenta_nombre(cuenta),
                        _format_currency(float(cuenta.get("saldo", 0))),
                    ]
                )
            equity_data.append(["", "UTILIDAD NETA", _format_currency(utilidad)])
            equity_data.append(
                ["", "TOTAL PATRIMONIO", _format_currency(patrimonio + utilidad)]
            )
            table = Table(equity_data, colWidths=[1.2 * inch, 2.3 * inch, 2 * inch])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BACKGROUND", (0, -2), (-1, -1), colors.HexColor("#e6f7e6")),
                        ("FONTNAME", (0, -2), (-1, -1), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -3),
                            [colors.white, colors.HexColor("#f9f9f9")],
                        ),
                    ]
                )
            )
            story.append(table)
        else:
            equity_data = [
                ["Concepto", "Valor (COP)"],
                ["Detalle no disponible", ""],
                ["Ganancias del Ejercicio", _format_currency(utilidad)],
                ["TOTAL PATRIMONIO", _format_currency(patrimonio + utilidad)],
            ]
            story.append(_build_table_2cols(equity_data, "#e6f7e6"))
        story.append(Spacer(1, 0.2 * inch))

        cuadre = report.get("cuadre", False)
        msg = report.get("mensaje_cuadre", "Balance not validated")
        color_cuadre = (
            colors.HexColor("#28a745") if cuadre else colors.HexColor("#dc3545")
        )
        story.append(
            Paragraph(
                f"<b>Validacion de Cuadre:</b> {_escape_paragraph_text(msg)}",
                ParagraphStyle(
                    "Balance", parent=styles["Normal"], textColor=color_cuadre
                ),
            )
        )

        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def to_excel(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance"

        header_fill = PatternFill(
            start_color="1F4788", end_color="1F4788", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        ws["A1"] = "BALANCE GENERAL"
        ws["A1"].font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells("A1:C1")

        ws["A2"] = f"Empresa: {company_name}"
        ws["A3"] = f"Periodo al: {report.get('period_end', '--')}"
        ws["A4"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        row = 6

        def write_section(title: str, items: list, total: float):
            nonlocal row
            ws[f"A{row}"] = title
            ws[f"A{row}"].font = header_font
            ws[f"A{row}"].fill = header_fill
            ws[f"B{row}"].fill = header_fill
            ws[f"C{row}"].fill = header_fill
            row += 1

            ws[f"A{row}"] = "Cuenta PUC"
            ws[f"B{row}"] = "Descripcion"
            ws[f"C{row}"] = "Valor"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(bold=True)
            ws[f"C{row}"].font = Font(bold=True)
            row += 1

            for cuenta in items:
                ws[f"A{row}"] = _get_cuenta_codigo(cuenta)
                ws[f"B{row}"] = _get_cuenta_nombre(cuenta)
                ws[f"C{row}"] = float(cuenta.get("saldo", 0))
                ws[f"C{row}"].number_format = "#,##0.00"
                row += 1

            ws[f"B{row}"] = f"TOTAL {title}"
            ws[f"C{row}"] = total
            ws[f"B{row}"].font = Font(bold=True)
            ws[f"C{row}"].font = Font(bold=True)
            ws[f"C{row}"].number_format = "#,##0.00"
            row += 2

        write_section(
            "ACTIVOS",
            report.get("activos_detalle") or [],
            float(report.get("activos", 0)),
        )
        write_section(
            "PASIVOS",
            report.get("pasivos_detalle") or [],
            float(report.get("pasivos", 0)),
        )

        patrimonio_items = list(report.get("patrimonio_detalle") or [])
        patrimonio_items.append(
            {
                "codigo": "",
                "nombre": "UTILIDAD NETA",
                "saldo": float(report.get("utilidad_neta", 0)),
            }
        )
        write_section(
            "PATRIMONIO",
            patrimonio_items,
            float(report.get("patrimonio", 0)) + float(report.get("utilidad_neta", 0)),
        )

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 20

        for c in ws["C"]:
            if isinstance(c.value, (int, float)):
                c.alignment = Alignment(horizontal="right")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class PnLExporter:
    """Export Profit & Loss reports to PDF and Excel."""

    @staticmethod
    def to_pdf(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#1f4788"),
            alignment=1,
        )
        section_style = ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontSize=11,
            textColor=colors.HexColor("#1f4788"),
            spaceBefore=8,
            spaceAfter=6,
        )

        story = []
        story.append(Paragraph("ESTADO DE RESULTADOS", title_style))
        story.append(
            Paragraph(
                f"<b>Empresa:</b> {_escape_paragraph_text(company_name)} | "
                f"<b>Periodo:</b> {_escape_paragraph_text(report.get('period_start', '--'))} "
                f"a {_escape_paragraph_text(report.get('period_end', '--'))}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.2 * inch))

        ingresos_total = float(report.get("total_ingresos", 0))
        story.append(Paragraph("INGRESOS OPERACIONALES", section_style))

        ingresos_data = [["Concepto", "Valor (COP)", "% Ingresos"]]
        for cuenta in report.get("ingresos", []):
            saldo = float(cuenta.get("saldo", 0))
            pct = (saldo / ingresos_total * 100) if ingresos_total > 0 else 0
            ingresos_data.append(
                [
                    f"{_get_cuenta_codigo(cuenta)} - {_get_cuenta_nombre(cuenta)}",
                    _format_currency(saldo),
                    f"{pct:.1f}%",
                ]
            )
        total_ingresos_pct = "100.0%" if ingresos_total > 0 else "0.0%"
        ingresos_data.append(
            ["TOTAL INGRESOS", _format_currency(ingresos_total), total_ingresos_pct]
        )

        ingresos_table = Table(
            ingresos_data, colWidths=[3 * inch, 1.8 * inch, 1.2 * inch]
        )
        ingresos_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8f0f7")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        story.append(ingresos_table)
        story.append(Spacer(1, 0.15 * inch))

        costo_total = float(report.get("total_costo_ventas", 0))
        story.append(Paragraph("COSTO DE VENTAS", section_style))

        costo_data = [["Concepto", "Valor (COP)"]]
        for cuenta in report.get("costo_ventas", []):
            costo_data.append(
                [
                    f"{_get_cuenta_codigo(cuenta)} - {_get_cuenta_nombre(cuenta)}",
                    _format_currency(float(cuenta.get("saldo", 0))),
                ]
            )
        costo_data.append(["TOTAL COSTO DE VENTAS", _format_currency(costo_total)])

        costo_table = Table(costo_data, colWidths=[4 * inch, 2 * inch])
        costo_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fff4e6")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        story.append(costo_table)
        story.append(Spacer(1, 0.15 * inch))

        utilidad_bruta = float(report.get("utilidad_bruta", 0))
        gastos_total = float(report.get("total_gastos", 0))
        utilidad_neta = float(report.get("utilidad_neta", 0))

        summary_data = [
            ["UTILIDAD BRUTA", _format_currency(utilidad_bruta)],
            ["GASTOS OPERACIONALES", _format_currency(gastos_total)],
            ["UTILIDAD NETA DEL EJERCICIO", _format_currency(utilidad_neta)],
        ]
        summary_table = Table(summary_data, colWidths=[4 * inch, 2 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8f0f7")),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e6f7e6")),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        story.append(summary_table)

        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def to_excel(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "P&L"

        header_fill = PatternFill(
            start_color="1F4788", end_color="1F4788", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        ws["A1"] = "ESTADO DE RESULTADOS"
        ws["A1"].font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells("A1:B1")

        ws["A2"] = f"Empresa: {company_name}"
        ws["A3"] = (
            f"Periodo: {report.get('period_start', '--')} - {report.get('period_end', '--')}"
        )

        row = 5
        ws[f"A{row}"] = "INGRESOS OPERACIONALES"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws[f"B{row}"].fill = header_fill
        row += 1

        for cuenta in report.get("ingresos", []):
            ws[f"A{row}"] = (
                f"{_get_cuenta_codigo(cuenta)} - {_get_cuenta_nombre(cuenta)}"
            )
            ws[f"B{row}"] = float(cuenta.get("saldo", 0))
            ws[f"B{row}"].number_format = "#,##0.00"
            row += 1

        ws[f"A{row}"] = "TOTAL INGRESOS"
        ws[f"B{row}"] = float(report.get("total_ingresos", 0))
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].number_format = "#,##0.00"
        row += 2

        ws[f"A{row}"] = "GASTOS OPERACIONALES"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws[f"B{row}"].fill = header_fill
        row += 1

        for cuenta in report.get("gastos", []):
            ws[f"A{row}"] = (
                f"{_get_cuenta_codigo(cuenta)} - {_get_cuenta_nombre(cuenta)}"
            )
            ws[f"B{row}"] = float(cuenta.get("saldo", 0))
            ws[f"B{row}"].number_format = "#,##0.00"
            row += 1

        ws[f"A{row}"] = "TOTAL GASTOS"
        ws[f"B{row}"] = float(report.get("total_gastos", 0))
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].number_format = "#,##0.00"
        row += 2

        ws[f"A{row}"] = "UTILIDAD NETA"
        ws[f"B{row}"] = float(report.get("utilidad_neta", 0))
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"B{row}"].font = Font(bold=True, size=12)
        ws[f"A{row}"].fill = PatternFill(
            start_color="E6F7E6", end_color="E6F7E6", fill_type="solid"
        )
        ws[f"B{row}"].fill = PatternFill(
            start_color="E6F7E6", end_color="E6F7E6", fill_type="solid"
        )
        ws[f"B{row}"].number_format = "#,##0.00"

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class CashFlowExporter:
    """Export Cash Flow reports to PDF and Excel."""

    @staticmethod
    def to_pdf(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#1f4788"),
            alignment=1,
        )
        section_style = ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontSize=11,
            textColor=colors.HexColor("#1f4788"),
        )

        story = []
        story.append(Paragraph("FLUJO DE CAJA", title_style))
        story.append(
            Paragraph(
                f"<b>Empresa:</b> {_escape_paragraph_text(company_name)} | "
                f"<b>Periodo:</b> {_escape_paragraph_text(report.get('period_start', '--'))} "
                f"a {_escape_paragraph_text(report.get('period_end', '--'))} | "
                f"<b>Metodo:</b> Directo",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.2 * inch))

        story.append(Paragraph("CUENTAS DE EFECTIVO Y BANCOS", section_style))
        data = [["Cuenta PUC", "Descripcion", "Saldo Neto (COP)"]]

        total_efectivo = 0.0
        for cuenta in report.get("cuentas_efectivo", []):
            saldo = float(cuenta.get("saldo", 0))
            total_efectivo += saldo
            data.append(
                [
                    _get_cuenta_codigo(cuenta),
                    _get_cuenta_nombre(cuenta),
                    _format_currency(saldo),
                ]
            )

        data.append(
            ["", "TOTAL EFECTIVO Y EQUIVALENTES", _format_currency(total_efectivo)]
        )

        table = Table(data, colWidths=[1.2 * inch, 2.8 * inch, 2 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8f0f7")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )

        story.append(table)
        story.append(Spacer(1, 0.2 * inch))
        story.append(
            Paragraph(
                f"<b>Metodologia:</b> {_escape_paragraph_text(report.get('nota', ''))}",
                styles["Normal"],
            )
        )

        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def to_excel(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        wb = Workbook()
        ws = wb.active
        ws.title = "Flujo de Caja"

        header_fill = PatternFill(
            start_color="1F4788", end_color="1F4788", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        ws["A1"] = "FLUJO DE CAJA - METODO DIRECTO"
        ws["A1"].font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells("A1:C1")

        ws["A2"] = f"Empresa: {company_name}"
        ws["A3"] = (
            f"Periodo: {report.get('period_start', '--')} - {report.get('period_end', '--')}"
        )

        row = 5
        headers = ["Cuenta PUC", "Descripcion", "Saldo (COP)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        total_efectivo = 0.0
        for cuenta in report.get("cuentas_efectivo", []):
            saldo = float(cuenta.get("saldo", 0))
            total_efectivo += saldo

            ws[f"A{row}"] = _get_cuenta_codigo(cuenta)
            ws[f"B{row}"] = _get_cuenta_nombre(cuenta)
            ws[f"C{row}"] = saldo
            ws[f"C{row}"].number_format = "#,##0.00"
            row += 1

        ws[f"A{row}"] = "TOTAL"
        ws[f"C{row}"] = total_efectivo
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"C{row}"].font = Font(bold=True, size=12)
        ws[f"A{row}"].fill = PatternFill(
            start_color="E8F0F7", end_color="E8F0F7", fill_type="solid"
        )
        ws[f"C{row}"].fill = PatternFill(
            start_color="E8F0F7", end_color="E8F0F7", fill_type="solid"
        )
        ws[f"C{row}"].number_format = "#,##0.00"

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class LibroDiarioExporter:
    """Export Libro Diario (Journal) to PDF and Excel."""

    @staticmethod
    def to_pdf(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor("#1f4788"),
        )

        story = []
        story.append(Paragraph("LIBRO DIARIO", title_style))
        story.append(
            Paragraph(
                f"<b>Empresa:</b> {_escape_paragraph_text(company_name)} | "
                f"<b>Periodo:</b> {_escape_paragraph_text(report.get('period_start', '--'))} "
                f"a {_escape_paragraph_text(report.get('period_end', '--'))}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.2 * inch))

        data = [["Fecha", "Cuenta", "Nombre Cuenta", "Descripcion", "Debito", "Credito"]]
        total_debito = 0.0
        total_credito = 0.0

        # Support builder key ("transacciones") and stored-statement key ("asientos").
        # Stored asientos are grouped vouchers; each may have nested "lineas" or "cuentas".
        raw_entries = report.get("transacciones") or report.get("asientos") or []

        def _iter_lines(entries):
            """Flatten grouped asientos into individual accounting lines."""
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                # Flat line (from live pipeline builder)
                if entry.get("cuenta_puc") or entry.get("debito") or entry.get("credito"):
                    yield entry
                    continue
                # Grouped voucher (from stored LibroDiarioContent)
                fecha_grupo = entry.get("fecha") or entry.get("fecha_comprobante") or ""
                comp = entry.get("comprobante") or entry.get("numero_comprobante") or ""
                desc_grupo = entry.get("descripcion") or entry.get("descripcion_general") or ""
                for linea in entry.get("lineas") or entry.get("cuentas") or entry.get("detalle") or []:
                    if isinstance(linea, dict):
                        yield {
                            "fecha": linea.get("fecha") or fecha_grupo,
                            "comprobante": linea.get("comprobante") or comp,
                            "cuenta_puc": linea.get("cuenta_puc") or linea.get("codigo_cuenta") or "",
                            "cuenta_nombre": linea.get("nombre_cuenta") or linea.get("cuenta_nombre") or "",
                            "descripcion": linea.get("descripcion") or desc_grupo,
                            "debito": linea.get("debito") or 0,
                            "credito": linea.get("credito") or 0,
                        }

        for trans in _iter_lines(raw_entries):
            debito = float(trans.get("debito") or 0)
            credito = float(trans.get("credito") or 0)
            total_debito += debito
            total_credito += credito

            fecha_raw = trans.get("fecha", "--") or "--"
            fecha = fecha_raw[:10] if fecha_raw and len(fecha_raw) > 10 else fecha_raw

            data.append(
                [
                    fecha,
                    trans.get("cuenta_puc", ""),
                    (trans.get("cuenta_nombre", "") or trans.get("descripcion", ""))[:25],
                    (trans.get("descripcion", "") or "")[:25],
                    _format_currency(debito),
                    _format_currency(credito),
                ]
            )

        data.append(["", "", "", "TOTAL", _format_currency(total_debito), _format_currency(total_credito)])

        table = Table(
            data,
            colWidths=[0.85 * inch, 0.75 * inch, 1.5 * inch, 1.5 * inch, 1.05 * inch, 1.05 * inch],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8f0f7")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f9f9f9")]),
                ]
            )
        )
        story.append(table)
        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def to_excel(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro Diario"

        header_fill = PatternFill(
            start_color="1F4788", end_color="1F4788", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        ws["A1"] = "LIBRO DIARIO"
        ws["A1"].font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells("A1:F1")

        ws["A2"] = f"Empresa: {company_name}"
        ws["A3"] = (
            f"Periodo: {report.get('period_start', '--')} - {report.get('period_end', '--')}"
        )

        row = 5
        headers = ["Fecha", "Cuenta PUC", "Nombre Cuenta", "Descripcion", "Debito", "Credito"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        total_debito = 0.0
        total_credito = 0.0

        raw_entries = report.get("transacciones") or report.get("asientos") or []

        def _iter_lines_xl(entries):
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                if entry.get("cuenta_puc") or entry.get("debito") or entry.get("credito"):
                    yield entry
                    continue
                fecha_g = entry.get("fecha") or entry.get("fecha_comprobante") or ""
                comp = entry.get("comprobante") or entry.get("numero_comprobante") or ""
                desc_g = entry.get("descripcion") or entry.get("descripcion_general") or ""
                for linea in entry.get("lineas") or entry.get("cuentas") or entry.get("detalle") or []:
                    if isinstance(linea, dict):
                        yield {
                            "fecha": linea.get("fecha") or fecha_g,
                            "comprobante": linea.get("comprobante") or comp,
                            "cuenta_puc": linea.get("cuenta_puc") or linea.get("codigo_cuenta") or "",
                            "cuenta_nombre": linea.get("nombre_cuenta") or linea.get("cuenta_nombre") or "",
                            "descripcion": linea.get("descripcion") or desc_g,
                            "debito": linea.get("debito") or 0,
                            "credito": linea.get("credito") or 0,
                        }

        for trans in _iter_lines_xl(raw_entries):
            debito = float(trans.get("debito") or 0)
            credito = float(trans.get("credito") or 0)
            total_debito += debito
            total_credito += credito

            fecha_raw = trans.get("fecha", "--") or "--"
            fecha = fecha_raw[:10] if fecha_raw and len(fecha_raw) > 10 else fecha_raw

            ws[f"A{row}"] = fecha
            ws[f"B{row}"] = trans.get("cuenta_puc", "")
            ws[f"C{row}"] = (trans.get("cuenta_nombre", "") or "")[:40]
            ws[f"D{row}"] = (trans.get("descripcion", "") or "")[:50]
            ws[f"E{row}"] = debito
            ws[f"F{row}"] = credito
            ws[f"E{row}"].number_format = "#,##0.00"
            ws[f"F{row}"].number_format = "#,##0.00"
            row += 1

        ws[f"A{row}"] = "TOTALES"
        ws[f"E{row}"] = total_debito
        ws[f"F{row}"] = total_credito
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"E{row}"].font = Font(bold=True, size=12)
        ws[f"F{row}"].font = Font(bold=True, size=12)
        ws[f"E{row}"].number_format = "#,##0.00"
        ws[f"F{row}"].number_format = "#,##0.00"

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class LibroAuxiliarExporter:
    """Export Libro Auxiliar (Subsidiary Ledger) to PDF and Excel."""

    @staticmethod
    def to_pdf(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor("#1f4788"),
        )
        section_style = ParagraphStyle(
            "Section", parent=styles["Heading2"], fontSize=10, textColor=colors.HexColor("#1f4788")
        )

        story = []
        story.append(Paragraph("LIBRO AUXILIAR", title_style))
        story.append(
            Paragraph(
                f"<b>Empresa:</b> {_escape_paragraph_text(company_name)} | "
                f"<b>Periodo:</b> {_escape_paragraph_text(report.get('period_start', '--'))} "
                f"a {_escape_paragraph_text(report.get('period_end', '--'))}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.2 * inch))

        for cuenta in report.get("cuentas", []):
            story.append(
                Paragraph(
                    f"Cuenta {cuenta.get('cuenta')} - {_escape_paragraph_text(cuenta.get('nombre', ''))[:40]}",
                    section_style,
                )
            )

            data = [["Fecha", "Descripcion", "Debito", "Credito", "Saldo"]]
            saldo = 0.0
            movimientos = cuenta.get("movimientos", []) or []

            for mov in movimientos:
                debito = float(mov.get("debito") or 0)
                credito = float(mov.get("credito") or 0)
                saldo += debito - credito

                fecha_raw = mov.get("fecha", "--") or "--"
                fecha = fecha_raw[:10] if fecha_raw and len(fecha_raw) > 10 else fecha_raw

                data.append(
                    [
                        fecha,
                        mov.get("descripcion", "")[:20],
                        _format_currency(debito),
                        _format_currency(credito),
                        _format_currency(saldo),
                    ]
                )

            table = Table(data, colWidths=[0.8 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1 * inch])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]
                )
            )
            story.append(table)
            story.append(Spacer(1, 0.15 * inch))

        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def to_excel(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro Auxiliar"

        header_fill = PatternFill(
            start_color="1F4788", end_color="1F4788", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=10)

        ws["A1"] = "LIBRO AUXILIAR"
        ws["A1"].font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells("A1:F1")

        ws["A2"] = f"Empresa: {company_name}"
        ws["A3"] = (
            f"Periodo: {report.get('period_start', '--')} - {report.get('period_end', '--')}"
        )

        row = 5
        for cuenta in report.get("cuentas", []):
            ws[f"A{row}"] = f"Cuenta {cuenta.get('cuenta')} - {cuenta.get('nombre', '')}"
            ws[f"A{row}"].font = Font(bold=True, size=11)
            ws.merge_cells(f"A{row}:F{row}")
            row += 1

            headers = ["Fecha", "Descripcion", "Debito", "Credito", "Saldo", ""]
            for col, header in enumerate(headers, 1):
                if header:
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
            row += 1

            saldo = 0.0
            for mov in cuenta.get("movimientos", []):
                debito = float(mov.get("debito") or 0)
                credito = float(mov.get("credito") or 0)
                saldo += debito - credito

                fecha_raw = mov.get("fecha", "--") or "--"
                fecha = fecha_raw[:10] if fecha_raw and len(fecha_raw) > 10 else fecha_raw
                ws[f"A{row}"] = fecha
                ws[f"B{row}"] = (mov.get("descripcion", "") or "")[:40]
                ws[f"C{row}"] = debito
                ws[f"D{row}"] = credito
                ws[f"E{row}"] = saldo
                ws[f"C{row}"].number_format = "#,##0.00"
                ws[f"D{row}"].number_format = "#,##0.00"
                ws[f"E{row}"].number_format = "#,##0.00"
                row += 1

            ws[f"A{row}"] = "TOTAL CUENTA"
            ws[f"C{row}"] = cuenta.get("total_debito", 0)
            ws[f"D{row}"] = cuenta.get("total_credito", 0)
            ws[f"E{row}"] = cuenta.get("saldo", 0)
            ws[f"A{row}"].font = Font(bold=True, size=10)
            ws[f"C{row}"].font = Font(bold=True, size=10)
            ws[f"D{row}"].font = Font(bold=True, size=10)
            ws[f"E{row}"].font = Font(bold=True, size=10)
            ws[f"C{row}"].number_format = "#,##0.00"
            ws[f"D{row}"].number_format = "#,##0.00"
            ws[f"E{row}"].number_format = "#,##0.00"
            row += 2

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class CambiosPatrimonioExporter:
    """Export Cambios en el Patrimonio (Equity Changes) to PDF and Excel."""

    @staticmethod
    def to_pdf(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor("#1f4788"),
        )

        story = []
        story.append(Paragraph("ESTADO DE CAMBIOS EN EL PATRIMONIO", title_style))
        story.append(
            Paragraph(
                f"<b>Empresa:</b> {_escape_paragraph_text(company_name)} | "
                f"<b>Periodo:</b> {_escape_paragraph_text(report.get('period_start', '--'))} "
                f"a {_escape_paragraph_text(report.get('period_end', '--'))}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.2 * inch))

        data = [["Cuenta PUC", "Descripcion", "Mov. Debito", "Mov. Credito", "Saldo Final"]]

        for cambio in report.get("cambios", []):
            data.append(
                [
                    cambio.get("codigo", ""),
                    _escape_paragraph_text(cambio.get("nombre", ""))[:30],
                    _format_currency(cambio.get("movimiento_debito", 0)),
                    _format_currency(cambio.get("movimiento_credito", 0)),
                    _format_currency(cambio.get("saldo_final", 0)),
                ]
            )

        table = Table(
            data,
            colWidths=[0.9 * inch, 1.8 * inch, 1.3 * inch, 1.3 * inch, 1.3 * inch],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4788")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8f0f7")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f9f9f9")]),
                ]
            )
        )
        story.append(table)
        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def to_excel(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        wb = Workbook()
        ws = wb.active
        ws.title = "Cambios Patrimonio"

        header_fill = PatternFill(
            start_color="1F4788", end_color="1F4788", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        ws["A1"] = "ESTADO DE CAMBIOS EN EL PATRIMONIO"
        ws["A1"].font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells("A1:E1")

        ws["A2"] = f"Empresa: {company_name}"
        ws["A3"] = (
            f"Periodo: {report.get('period_start', '--')} - {report.get('period_end', '--')}"
        )

        row = 5
        headers = ["Cuenta PUC", "Descripcion", "Movimiento Debito", "Movimiento Credito", "Saldo Final"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        for cambio in report.get("cambios", []):
            ws[f"A{row}"] = cambio.get("codigo", "")
            ws[f"B{row}"] = cambio.get("nombre", "")[:40]
            ws[f"C{row}"] = float(cambio.get("movimiento_debito", 0))
            ws[f"D{row}"] = float(cambio.get("movimiento_credito", 0))
            ws[f"E{row}"] = float(cambio.get("saldo_final", 0))
            ws[f"C{row}"].number_format = "#,##0.00"
            ws[f"D{row}"].number_format = "#,##0.00"
            ws[f"E{row}"].number_format = "#,##0.00"
            row += 1

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class NotasEstadosFinancierosExporter:
    """Export Notas a los Estados Financieros (Notes) to PDF and Excel."""

    @staticmethod
    def to_pdf(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor("#1f4788"),
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=11,
            textColor=colors.HexColor("#1f4788"),
        )

        story = []
        story.append(Paragraph("NOTAS A LOS ESTADOS FINANCIEROS", title_style))
        story.append(
            Paragraph(
                f"<b>Empresa:</b> {_escape_paragraph_text(company_name)} | "
                f"<b>Periodo:</b> {_escape_paragraph_text(report.get('period_end', '--'))}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.2 * inch))

        # Resumen financiero
        resumen = report.get("resumen_financiero", {})
        story.append(Paragraph("<b>Resumen de Situacion Financiera:</b>", heading_style))
        data = [
            ["Activos", _format_currency(resumen.get("activos", 0))],
            ["Pasivos", _format_currency(resumen.get("pasivos", 0))],
            ["Patrimonio", _format_currency(resumen.get("patrimonio", 0))],
        ]
        table = _build_table_2cols(data, "#e8f0f7")
        story.append(table)
        story.append(Spacer(1, 0.2 * inch))

        # Notas normativas
        notas = report.get("notas", [])
        if notas:
            story.append(Paragraph("<b>Notas Normativas:</b>", heading_style))
            for nota_item in notas[:5]:
                numero = nota_item.get("numero", 0)
                titulo = nota_item.get("titulo", "")
                contenido = nota_item.get("contenido", "")
                story.append(
                    Paragraph(
                        f"<b>Nota {numero}: {_escape_paragraph_text(titulo)}</b>",
                        styles["Heading3"],
                    )
                )
                story.append(Paragraph(f"{_escape_paragraph_text(contenido)}", styles["Normal"]))
                story.append(Spacer(1, 0.1 * inch))
        else:
            story.append(Paragraph("No hay notas disponibles.", styles["Normal"]))

        doc.build(story)
        return buffer.getvalue()

    @staticmethod
    def to_excel(report: Dict[str, Any], company_name: str = "Empresa") -> bytes:
        if not report:
            raise ValueError("Report data is empty or None")
        wb = Workbook()
        ws = wb.active
        ws.title = "Notas"

        header_fill = PatternFill(
            start_color="1F4788", end_color="1F4788", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        ws["A1"] = "NOTAS A LOS ESTADOS FINANCIEROS"
        ws["A1"].font = Font(bold=True, size=14, color="1F4788")
        ws.merge_cells("A1:B1")

        ws["A2"] = f"Empresa: {company_name}"
        ws["A3"] = f"Periodo: {report.get('period_end', '--')}"

        row = 5
        ws["A5"] = "Resumen Financiero"
        ws["A5"].font = Font(bold=True, size=12)
        row = 6

        resumen = report.get("resumen_financiero", {})
        headers = ["Item", "Valor"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        for item, valor in [
            ("Activos", resumen.get("activos", 0)),
            ("Pasivos", resumen.get("pasivos", 0)),
            ("Patrimonio", resumen.get("patrimonio", 0)),
        ]:
            ws[f"A{row}"] = item
            ws[f"B{row}"] = valor
            ws[f"B{row}"].number_format = "#,##0.00"
            row += 1

        row += 2
        ws[f"A{row}"] = "Notas Normativas"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        notas = report.get("notas", [])
        for nota_item in notas[:10]:
            numero = nota_item.get("numero", 0)
            titulo = nota_item.get("titulo", "")
            contenido = nota_item.get("contenido", "")
            ws[f"A{row}"] = f"Nota {numero}: {titulo}"
            ws[f"A{row}"].font = Font(bold=True, size=10)
            row += 1
            ws[f"A{row}"] = contenido[:100]
            ws[f"A{row}"].alignment = Alignment(wrap_text=True)
            row += 2

        ws.column_dimensions["A"].width = 80
        ws.column_dimensions["B"].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
